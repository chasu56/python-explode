import openpyxl

# Load the original Excel file and sheet
workbook = openpyxl.load_workbook('original_file.xlsx')
sheet = workbook.active

# Get the data from the sheet and create a list of tuples
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    name, age, nationality, sex = row
    age = age.split(',') if isinstance(age, str) else ['']
    nationality = nationality.split(',') if isinstance(nationality, str) else ['']
    sex = sex.split(',') if isinstance(sex, str) else ['']
    max_len = max(len(age), len(nationality), len(sex))
    age += [''] * (max_len - len(age))
    nationality += [''] * (max_len - len(nationality))
    sex += [''] * (max_len - len(sex))
    data.append(tuple(zip(name.split(','), age, nationality, sex)))

# Explode the data into a list of dictionaries
exploded_data = []
for item in data:
    for i in range(len(item[0])):
        row = {
            'Name': item[0][i].strip(),
            'Age': item[1][i].strip(),
            'Nationality': item[2][i].strip(),
            'Sex': item[3][i].strip()
        }
        if all(row.values()):
            exploded_data.append(row)

# Write the exploded data to a new sheet
new_workbook = openpyxl.Workbook()
new_sheet = new_workbook.active
new_sheet.append(['Name', 'Age', 'Nationality', 'Sex'])
for item in exploded_data:
    new_sheet.append([item['Name'], item['Age'], item['Nationality'], item['Sex']])

# Save the new workbook to a new file
new_workbook.save('new_file.xlsx')
